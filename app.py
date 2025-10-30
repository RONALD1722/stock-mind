import os
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, send_file, flash
from werkzeug.utils import secure_filename
from config import conectar, desconectar
import bcrypt
from functools import wraps

from io import BytesIO

# librerias para exel 
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Extras de psycopg2
import psycopg2
from psycopg2.extras import RealDictCursor

# Librerías para correos
import smtplib
from email.mime.text import MIMEText

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "clave_super_secreta")

def requiere_rol(*roles_permitidos):
    def wrapper(f):
        @wraps(f)
        def decorado(*args, **kwargs):
            if "rol" not in session:
                flash("Debes iniciar sesión primero", "error")
                return redirect(url_for("login"))
            if session["rol"] not in roles_permitidos:
                flash("No tienes permiso para acceder a esta sección", "error")
                return redirect(url_for("inicio"))
            return f(*args, **kwargs)
        return decorado
    return wrapper


# Carpeta para subir imágenes
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Configuración SMTP (en producción usa variables de entorno)
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587))
EMAIL_USER = os.environ.get("EMAIL_USER", "StockMind21@gmail.com")
EMAIL_PASS = os.environ.get("EMAIL_PASS", "arnh yzvg ckii urcd")  # EN PRODUCCIÓN: NO dejar aquí.

# ---------- RUTA PARA SERVIR UPLOADS ----------
@app.route('/uploads/<path:filename>')
def uploads(filename):
    # seguridad: secure_filename no se aplica aquí porque ya fue guardado con secure_filename
    return send_from_directory(UPLOAD_FOLDER, filename)

# ---------- LOGIN ----------
@app.route("/StockMind/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        clave = request.form.get("clave", "")
        rol = request.form.get("tipo_usuario", "")

        conexion = conectar()
        if not conexion:
            return "No se pudo conectar a la base de datos", 500

        try:
            cur = conexion.cursor()
            cur.execute("""
                SELECT id_usu, nombre, usuario, rol, clave, ultimo_login, fecha_registro
                FROM usuarios
                WHERE usuario = %s AND rol = %s
            """, (usuario, rol))
            user = cur.fetchone()
            cur.close()
        except Exception as e:
            desconectar(conexion)
            return f"Error en consulta de login: {e}", 500
        finally:
            desconectar(conexion)

        if user:
            try:
                hash_db = user[4].encode('utf-8') if isinstance(user[4], str) else user[4]
                if bcrypt.checkpw(clave.encode('utf-8'), hash_db):

                    # Consultar préstamos activos del usuario
                    conexion = conectar()
                    prestamos_usuario_activos = 0
                    if conexion:
                        try:
                            cur = conexion.cursor()
                            cur.execute("""
                                SELECT COUNT(*) 
                                FROM prestamos 
                                WHERE id_usu = %s AND estado = 'pendiente'
                            """, (user[0],))
                            prestamos_usuario_activos = cur.fetchone()[0] or 0
                            cur.close()
                        except Exception:
                            prestamos_usuario_activos = 0
                        finally:
                            desconectar(conexion)

                    # Guardar datos en sesión
                    session.update({
                        'id_usuario': user[0],
                        'nombre': user[1],
                        'usuario': user[2],
                        'rol': user[3],
                        'ultimo_login': user[5].strftime("%d/%m/%Y %H:%M:%S") if user[5] else "Primera vez",
                        'fecha_registro': user[6].strftime("%d/%m/%Y %H:%M:%S") if user[6] else "No disponible",
                        'prestamos_usuario_activos': prestamos_usuario_activos   # 👈 nombre único
                    })


                    # actualizar ultimo login a la fecha actual
                    conexion = conectar()
                    if conexion:
                        try:
                            cur = conexion.cursor()
                            cur.execute("UPDATE usuarios SET ultimo_login = CURRENT_TIMESTAMP WHERE id_usu=%s", (user[0],))
                            conexion.commit()
                            cur.close()
                        except Exception:
                            pass
                        finally:
                            desconectar(conexion)

                    return redirect(url_for("inicio"))
                else:
                    return render_template("login.html", error="Contraseña incorrecta")
            except Exception as e:
                return render_template("login.html", error=f"Error validando contraseña: {e}")
        else:
            return render_template("login.html", error="Usuario no encontrado")

    return render_template("login.html")


# ---------- INDEX ----------
@app.route("/StockMind/index")
@requiere_rol("administrador", "encargado")
def index():
    if "usuario" not in session:
        return redirect(url_for("login"))
    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    productos = []
    try:
        cur = conexion.cursor()
        cur.execute("""
            SELECT 
                p.id_prod,
                p.nombre,
                p.cantidad AS cantidad_total,
                p.lugar,
                p.descripcion,
                p.imagen,
                p.activo,
                COALESCE(SUM(CASE WHEN pr.estado = 'pendiente' THEN pr.cantidad ELSE 0 END), 0) AS prestados,
                p.cantidad - COALESCE(SUM(CASE WHEN pr.estado = 'pendiente' THEN pr.cantidad ELSE 0 END), 0) AS disponible
            FROM producto p
            LEFT JOIN prestamos pr ON p.id_prod = pr.id_prod
            GROUP BY 
                p.id_prod, p.nombre, p.cantidad, p.lugar, p.descripcion, p.imagen, p.activo
            ORDER BY p.id_prod DESC;
        """)

        rows = cur.fetchall()
        for r in rows:
            productos.append({
                "id_prod": r[0],
                "nombre": r[1],
                "cantidad_total": int(r[2]) if r[2] is not None else 0,
                "lugar": r[3],
                "descripcion": r[4],
                "imagen": r[5],
                "activo": r[6],
                "prestados": int(r[7]) if r[7] is not None else 0,
                "disponible": int(r[8]) if r[8] is not None else 0
            })

        cur.close()
    except Exception as e:
        return f"Error consultando productos: {e}", 500
    finally:
        desconectar(conexion)

    return render_template("index.html", productos=productos, active_page="index")


#---------- FUNCIÓN PARA ENVIAR CORREOS ----------
def enviar_correo(destinatario, asunto, cuerpo_html):
    if not destinatario:
        print("No hay destinatario válido para el correo.")
        return
    msg = MIMEText(cuerpo_html, "html")
    msg["Subject"] = asunto
    msg["From"] = EMAIL_USER
    msg["To"] = destinatario

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(msg["From"], [msg["To"]], msg.as_string())
        print(f"✅ Correo enviado a {destinatario}")
    except Exception as e:
        print(f"Error enviando correo a {destinatario}: {e}")


# ---------- PLANTILLA DE CORREO CORPORATIVA ----------
def plantilla_correo(titulo, mensaje, usuario, producto, cantidad, fecha=None, extra="", color="#004080"):
    logo_url = "/static/logo.png.png"  # Cambia por el logo oficial si lo tienes

    cuerpo = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: auto; 
                border: 1px solid #ddd; border-radius: 10px; overflow: hidden;">
        <div style="background: {color}; color: white; padding: 20px; text-align: center;">
            <img src="{logo_url}" alt="StockMind" style="max-width: 120px; margin-bottom: 10px;">
            <h2 style="margin: 0;">{titulo}</h2>
        </div>
        <div style="padding: 20px; color: #333;">
            <p>Hola <b>{usuario}</b>,</p>
            <p>{mensaje}</p>
            
            <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                <tr>
                    <td style="padding: 10px; border: 1px solid #ccc; background:#f9f9f9;"><b>Producto</b></td>
                    <td style="padding: 10px; border: 1px solid #ccc;">{producto}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ccc; background:#f9f9f9;"><b>Cantidad</b></td>
                    <td style="padding: 10px; border: 1px solid #ccc;">{cantidad}</td>
                </tr>
                {f"<tr><td style='padding:10px; border:1px solid #ccc; background:#f9f9f9;'><b>Fecha</b></td><td style='padding:10px; border:1px solid #ccc;'>{fecha}</td></tr>" if fecha else ""}
            </table>

            {extra}
        </div>
        <div style="background: #f4f4f4; padding: 15px; text-align: center; font-size: 12px; color: #777;">
            <p>StockMind © 2025 - Sistema de gestión de inventario</p>
        </div>
    </div>
    """
    return cuerpo


# ---------- NOTIFICACIÓN DE PRÉSTAMO ----------
def notificacion_prestamo(id_prestamo):
    try:
        conn = psycopg2.connect(dbname="inventario3", user="postgres", password="123456",
                                host="localhost", port="5432")
        cur = conn.cursor()
        cur.execute("""
            SELECT u.correo, u.nombre, p.nombre, pr.cantidad, pr.fecha_devolucion
            FROM prestamos pr
            JOIN usuarios u ON pr.id_usu = u.id_usu
            JOIN producto p ON pr.id_prod = p.id_prod
            WHERE pr.id_prestamo = %s;
        """, (id_prestamo,))
        fila = cur.fetchone()
        if not fila:
            cur.close()
            conn.close()
            return
        correo, usuario, producto, cantidad, fecha_dev = fila
        dias_restantes = (fecha_dev.date() - date.today()).days

        extra = f"<p><b>Días restantes:</b> {dias_restantes} día(s)</p>"

        cuerpo = plantilla_correo(
            titulo="📦 Confirmación de préstamo",
            mensaje=f"Se ha registrado tu préstamo en <b>StockMind</b>.",
            usuario=usuario,
            producto=producto,
            cantidad=cantidad,
            fecha=fecha_dev.date(),
            extra=extra
        )

        enviar_correo(correo, "Confirmación de tu préstamo 📌", cuerpo)
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error notificacion_prestamo: {e}")


def notificacion_devolucion(id_prestamo):
    try:
        conn = psycopg2.connect(
            dbname="inventario3", user="postgres", password="123456",
            host="localhost", port="5432"
        )
        cur = conn.cursor()

        # Datos del préstamo (correo, usuario y producto)
        cur.execute("""
            SELECT u.correo, u.nombre, p.nombre
            FROM prestamos pr
            JOIN usuarios u ON pr.id_usu = u.id_usu
            JOIN producto p ON pr.id_prod = p.id_prod
            WHERE pr.id_prestamo = %s;
        """, (id_prestamo,))
        fila = cur.fetchone()
        if not fila:
            cur.close()
            conn.close()
            return
        correo, usuario, producto = fila

        # Última devolución registrada en historial
        cur.execute("""
            SELECT detalle, fecha
            FROM historial
            WHERE id_prod = (SELECT id_prod FROM prestamos WHERE id_prestamo=%s)
              AND id_usu = (SELECT id_usu FROM prestamos WHERE id_prestamo=%s)
              AND accion = 'Devolución'
            ORDER BY fecha DESC
            LIMIT 1;
        """, (id_prestamo, id_prestamo))
        historial = cur.fetchone()

        cantidad = "desconocida"
        fecha_real = None
        if historial:
            detalle, fecha_real = historial
            if "Devolución de" in detalle:
                try:
                    cantidad = int(detalle.split("Devolución de ")[1].split(" ")[0])
                except:
                    pass

        extra = "<p>¡Gracias por cumplir con los tiempos del préstamo! 🎉</p>"

        cuerpo = plantilla_correo(
            titulo="✅ Confirmación de devolución",
            mensaje="Hemos registrado la devolución de tu préstamo en <b>StockMind</b>.",
            usuario=usuario,
            producto=producto,
            cantidad=cantidad,
            fecha=fecha_real.date() if fecha_real else None,
            extra=extra
        )

        enviar_correo(correo, "Confirmación de devolución ✅", cuerpo)
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error notificacion_devolucion: {e}")




# ---------- NOTIFICACIÓN DE RECORDATORIO ----------
def notificacion_recordatorios():
    try:
        conn = psycopg2.connect(dbname="inventario3", user="postgres", password="123456",
                                host="localhost", port="5432")
        cur = conn.cursor()
        cur.execute("""
            SELECT u.correo, u.nombre, p.nombre, pr.cantidad, pr.fecha_devolucion
            FROM prestamos pr
            JOIN usuarios u ON pr.id_usu = u.id_usu
            JOIN producto p ON pr.id_prod = p.id_prod
            WHERE pr.estado = 'pendiente';
        """)
        prestamos = cur.fetchall()
        for correo, usuario, producto, cantidad, fecha_dev in prestamos:
            dias_restantes = (fecha_dev.date() - date.today()).days
            if dias_restantes == 1:
                extra = """
                <div style="margin-top:20px; padding:10px; background:#fff3cd; 
                            border:1px solid #ffeeba; border-radius:5px; color:#856404;">
                    ⚠️ Recuerda que mañana vence el plazo para la devolución.
                </div>
                """
                cuerpo = plantilla_correo(
                    titulo="⏰ Recordatorio de devolución",
                    mensaje="Este es un recordatorio de <b>StockMind</b> para que realices la devolución a tiempo.",
                    usuario=usuario,
                    producto=producto,
                    cantidad=cantidad,
                    fecha=fecha_dev.date(),
                    extra=extra,
                    color="#e67e22"
                )
                enviar_correo(correo, "Recordatorio: entrega mañana 📌", cuerpo)
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error notificacion_recordatorios: {e}")

@app.route("/test_recordatorio_33")
def test_recordatorio_33():
    notificacion_recordatorios()
    return "✅ Recordatorio ejecutado para préstamo 33 (revisa tu correo)"



# ---------- AGREGAR / ACTUALIZAR SI EXISTE ----------
@app.route("/StockMind/agregar", methods=["POST"])
@requiere_rol("administrador","encargado")
def agregar():
    nombre = request.form.get("nombre", "").strip()
    cantidad = request.form.get("cantidad", "0")
    lugar = request.form.get("lugar", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    imagen = request.files.get("imagen")

    try:
        cantidad_int = int(cantidad)
    except ValueError:
        cantidad_int = 0

    imagen_nombre = "default_prod.png"
    if imagen and imagen.filename:
        imagen_nombre = secure_filename(imagen.filename)
        imagen.save(os.path.join(UPLOAD_FOLDER, imagen_nombre))

    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    id_prod = None
    try:
        cur = conexion.cursor()

        # Verificar si ya existe un producto con el mismo nombre
        cur.execute("SELECT id_prod, cantidad FROM producto WHERE LOWER(nombre) = LOWER(%s)", (nombre,))
        fila = cur.fetchone()

        if fila:
            # Ya existe → actualizar sumando cantidad
            id_prod = fila[0]
            cantidad_existente = fila[1] or 0
            nueva_cantidad = cantidad_existente + cantidad_int

            cur.execute("""
                UPDATE producto SET
                    cantidad = %s,
                    lugar = %s,
                    descripcion = %s,
                    imagen = %s,
                    activo = TRUE
                WHERE id_prod = %s
            """, (nueva_cantidad, lugar, descripcion, imagen_nombre, id_prod))

            accion = "Actualizar producto"
            detalle = f"Se actualizó el producto {nombre} (ID {id_prod}), cantidad {cantidad_existente} → {nueva_cantidad}"

        else:
            # No existe → insertar nuevo
            cur.execute("""
                INSERT INTO producto (nombre, cantidad, lugar, descripcion, imagen)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id_prod
            """, (nombre, cantidad_int, lugar, descripcion, imagen_nombre))
            row = cur.fetchone()
            id_prod = row[0] if row else None
            accion = "Registrar producto"
            detalle = f"Se registró el producto {nombre} (ID {id_prod}) con cantidad {cantidad_int}"

        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        return f"Error insertando/actualizando producto: {e}", 500
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if "id_usuario" in session and id_prod:
        registrar_historial(session["id_usuario"], accion, detalle)

    return redirect(url_for("index"))


@app.route("/StockMind/buscar_producto", methods=["GET"])
def buscar_producto():
    nombre = request.args.get("nombre", "").strip()
    if not nombre:
        return {"existe": False}

    conexion = conectar()
    if not conexion:
        return {"error": "No se pudo conectar a la base de datos"}, 500

    try:
        cur = conexion.cursor()
        cur.execute("""
            SELECT id_prod, nombre, cantidad, lugar, descripcion, imagen, activo
            FROM producto
            WHERE LOWER(nombre) = LOWER(%s)
        """, (nombre,))
        fila = cur.fetchone()
        cur.close()
    finally:
        desconectar(conexion)

    if fila:
        return {
            "existe": True,
            "id_prod": fila[0],
            "nombre": fila[1],
            "cantidad": fila[2],
            "lugar": fila[3],
            "descripcion": fila[4],
            "imagen": fila[5],
            "activo": fila[6]
        }
    return {"existe": False}



# ---------- ELIMINAR / RESTAURAR PRODUCTO ----------
@app.route("/StockMind/eliminar_producto/<int:id_prod>", methods=["POST"])
def eliminar_producto(id_prod):
    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500
    producto_nombre = None
    try:
        cur = conexion.cursor()
        cur.execute("SELECT nombre FROM producto WHERE id_prod=%s", (id_prod,))
        fila = cur.fetchone()
        producto_nombre = fila[0] if fila else None
        cur.execute("UPDATE producto SET activo = FALSE WHERE id_prod = %s", (id_prod,))
        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        return f"Error desactivando producto: {e}", 500
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if producto_nombre and "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Desactivar producto", f"Se desactivó el producto {producto_nombre}")

    return redirect(url_for("index"))


@app.route("/StockMind/restaurar_producto/<int:id_prod>", methods=["POST"])
def restaurar_producto(id_prod):
    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500
    producto_nombre = None
    try:
        cur = conexion.cursor()
        cur.execute("SELECT nombre FROM producto WHERE id_prod=%s", (id_prod,))
        fila = cur.fetchone()
        producto_nombre = fila[0] if fila else None
        cur.execute("UPDATE producto SET activo = TRUE WHERE id_prod = %s", (id_prod,))
        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        return f"Error restaurando producto: {e}", 500
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if producto_nombre and "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Restaurar producto", f"Se restauró el producto {producto_nombre}")

    return redirect(url_for("index"))


# ---------- ACTUALIZAR PRODUCTO ----------
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/actualizar_producto", methods=["POST"])
def actualizar_producto():
    id_prod = request.form.get("id_prod")
    nombre = request.form.get("nombre")
    cantidad = request.form.get("cantidad")
    lugar = request.form.get("lugar")
    descripcion = request.form.get("descripcion")
    imagen_file = request.files.get("imagen")

    if not id_prod:
        flash("ID de producto no proporcionado", "danger")
        return redirect(url_for("index"))

    try:
        cantidad_int = int(cantidad)
    except (ValueError, TypeError):
        cantidad_int = 0

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("index"))

    producto_actual = None
    try:
        cursor = conexion.cursor()

        # Obtener datos actuales
        cursor.execute("SELECT nombre, imagen FROM producto WHERE id_prod=%s", (id_prod,))
        fila = cursor.fetchone()
        producto_actual = fila[0] if fila else None
        imagen_actual = fila[1] if fila else None

        # Imagen: mantener la actual por defecto
        imagen_nombre = imagen_actual

        # Si se sube una nueva y es válida, reemplazarla
        if imagen_file and imagen_file.filename.strip() != "":
            if allowed_file(imagen_file.filename):
                filename = secure_filename(imagen_file.filename)
                ruta_imagen = os.path.join(UPLOAD_FOLDER, filename)
                imagen_file.save(ruta_imagen)
                imagen_nombre = filename
            else:
                flash("Formato de imagen no permitido", "warning")

        # Actualizar producto
        consulta = """
            UPDATE producto SET
            nombre = %s,
            cantidad = %s,
            lugar = %s,
            descripcion = %s,
            imagen = %s
            WHERE id_prod = %s
        """
        datos = (nombre, cantidad_int, lugar, descripcion, imagen_nombre, id_prod)
        cursor.execute(consulta, datos)
        conexion.commit()
        cursor.close()

        flash(f"Producto '{nombre}' actualizado correctamente", "success")

    except Exception as e:
        conexion.rollback()
        print(f"Error al actualizar producto: {e}")
        flash(f"Error al actualizar: {e}", "danger")
    finally:
        desconectar(conexion)

    # Historial
    if producto_actual and "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Editar producto", f"Se actualizó el producto {producto_actual}")

    return redirect(url_for("index"))




# ---------- LISTAR USUARIOS ----------
@app.route("/StockMind/usuarios")
@requiere_rol("administrador")
def listar_usuarios():
    if "usuario" not in session:
        return redirect(url_for("login"))
    conexion = conectar()
    usuarios = []
    if not conexion:
        return "No se pudo conectar a la base de datos", 500
    try:
        cur = conexion.cursor()
        cur.execute("SELECT id_usu, nombre, usuario, rol, telefono, correo, direccion, foto FROM usuarios ORDER BY id_usu")
        usuarios = [
            {"id_usu": r[0], "nombre": r[1], "usuario": r[2], "rol": r[3], "telefono": r[4], "correo": r[5], "direccion": r[6], "foto": r[7]}
            for r in cur.fetchall()
        ]
        cur.close()
    except Exception as e:
        return f"Error listando usuarios: {e}", 500
    finally:
        desconectar(conexion)

    return render_template("usuarios.html", usuarios=usuarios, active_page="usuarios")


@app.route("/StockMind/usuarios/agregar", methods=["POST"])
def agregar_usuario():
    nombre = request.form.get("nombre", "").strip()
    usuario = request.form.get("usuario", "").strip()
    rol = request.form.get("tipo_usuario", "").strip()
    clave = request.form.get("clave", "")
    telefono = request.form.get("telefono", "")
    correo = request.form.get("correo")
    direccion = request.form.get("direccion")
    foto = request.files.get("foto")

    if not usuario or not clave:
        flash("Usuario y contraseña son obligatorios", "danger")
        return redirect(url_for("listar_usuarios"))

    hashed = bcrypt.hashpw(clave.encode('utf-8'), bcrypt.gensalt())

    foto_nombre = "default.png"
    if foto and foto.filename:
        foto_nombre = secure_filename(f"{usuario}_{foto.filename}")
        foto.save(os.path.join(UPLOAD_FOLDER, foto_nombre))

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("listar_usuarios"))

    try:
        cur = conexion.cursor()
        cur.execute("""
            INSERT INTO usuarios (nombre, usuario, rol, clave, telefono, correo, direccion, foto)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (nombre, usuario, rol, hashed.decode('utf-8'), telefono, correo, direccion, foto_nombre))
        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        flash(f"Error agregando usuario: {e}", "danger")
        return redirect(url_for("listar_usuarios"))
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Agregar usuario", f"Se registró al usuario {usuario}")

    return redirect(url_for("listar_usuarios"))


@app.route("/StockMind/usuarios/editar", methods=["POST"])
def editar_usuario():
    id_usu_form = request.form.get("id_usu")
    if id_usu_form:
        id_usu = int(id_usu_form)
    else:
        id_usu = session.get("id_usuario")

    nombre = request.form.get("nombre")
    usuario = request.form.get("usuario")
    telefono = request.form.get("telefono")
    clave = request.form.get("clave")
    rol = request.form.get("tipo_usuario")
    correo = request.form.get("correo")
    direccion = request.form.get("direccion")
    foto = request.files.get("foto")

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("listar_usuarios"))
    try:
        cur = conexion.cursor()
        cur.execute("SELECT foto FROM usuarios WHERE id_usu=%s", (id_usu,))
        fila = cur.fetchone()
        foto_actual = fila[0] if fila and fila[0] else "default.png"
        foto_nombre = foto_actual

        if foto and foto.filename:
            foto_nombre = secure_filename(f"{id_usu}_{foto.filename}")
            foto.save(os.path.join(UPLOAD_FOLDER, foto_nombre))

        if clave:
            hashed = bcrypt.hashpw(clave.encode('utf-8'), bcrypt.gensalt())
            consulta = """
                UPDATE usuarios
                SET nombre=%s, usuario=%s, rol=%s, clave=%s,
                    telefono=%s, correo=%s, direccion=%s, foto=%s
                WHERE id_usu=%s
            """
            cur.execute(consulta, (nombre, usuario, rol, hashed.decode('utf-8'),
                                   telefono, correo, direccion, foto_nombre, id_usu))
        else:
            consulta = """
                UPDATE usuarios
                SET nombre=%s, usuario=%s, rol=%s,
                    telefono=%s, correo=%s, direccion=%s, foto=%s
                WHERE id_usu=%s
            """
            cur.execute(consulta, (nombre, usuario, rol, telefono,
                                   correo, direccion, foto_nombre, id_usu))

        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        flash(f"Error editando usuario: {e}", "danger")
        return redirect(url_for("listar_usuarios"))
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Editar usuario", f"Se editó al usuario {usuario}")

    if "id_usuario" in session and session["id_usuario"] == id_usu:
        return redirect(url_for("perfil"))

    return redirect(url_for("listar_usuarios"))


# ---------- ELIMINAR USUARIO ----------
@app.route("/StockMind/usuarios/eliminar/<int:id_usu>", methods=["POST"])
def eliminar_usuario(id_usu):
    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("listar_usuarios"))
    usuario_eliminado = None
    try:
        cur = conexion.cursor()
        cur.execute("SELECT usuario FROM usuarios WHERE id_usu=%s", (id_usu,))
        fila = cur.fetchone()
        usuario_eliminado = fila[0] if fila else None
        cur.execute("DELETE FROM usuarios WHERE id_usu=%s", (id_usu,))
        conexion.commit()
        cur.close()
    except Exception as e:
        conexion.rollback()
        flash(f"Error eliminando usuario: {e}", "danger")
        return redirect(url_for("listar_usuarios"))
    finally:
        desconectar(conexion)

    # --- Registrar en historial ---
    if usuario_eliminado and "id_usuario" in session:
        registrar_historial(session["id_usuario"], "Eliminar usuario", f"Se eliminó al usuario {usuario_eliminado}")

    return redirect(url_for("listar_usuarios"))


# ---------- PERFIL ----------
@app.route("/StockMind/perfil")
@requiere_rol("administrador","usuario", "encargado")
def perfil():
    if "id_usuario" not in session:
        return redirect(url_for("login"))

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("index"))

    usuario = None
    prestamos_pendientes = []
    try:
        cur = conexion.cursor()

        # Datos del usuario
        cur.execute("""
            SELECT id_usu, nombre, usuario, rol, telefono, correo, direccion, ultimo_login, foto
            FROM usuarios
            WHERE id_usu=%s
        """, (session["id_usuario"],))
        row = cur.fetchone()
        if row:
            usuario = {
                "id": row[0], "nombre": row[1], "usuario": row[2], "rol": row[3],
                "telefono": row[4], "correo": row[5], "direccion": row[6],
                "ultimo_login": row[7], "foto": row[8]
            }

        # Préstamos pendientes del usuario
        cur.execute("""
            SELECT pr.id_prestamo, p.nombre AS producto, pr.cantidad, 
                pr.fecha_prestamo, pr.fecha_devolucion, pr.estado
            FROM prestamos pr
            JOIN producto p ON pr.id_prod = p.id_prod
            WHERE pr.id_usu = %s AND pr.estado = 'pendiente'
            ORDER BY pr.fecha_devolucion ASC
        """, (session["id_usuario"],))

        prestamos_pendientes = []
        for row in cur.fetchall():
            id_prestamo, producto, cantidad, fecha_prestamo, fecha_devolucion, estado = row
            prestamos_pendientes.append((
                id_prestamo,
                producto,
                cantidad,
                fecha_prestamo.date() if hasattr(fecha_prestamo, "date") else fecha_prestamo,
                fecha_devolucion.date() if hasattr(fecha_devolucion, "date") else fecha_devolucion,
                estado
            ))

        cur.close()
    except Exception as e:
        flash(f"Error cargando perfil: {e}", "danger")
    finally:
        desconectar(conexion)

    # 🔹 Pasamos también la fecha de hoy
    return render_template("perfil_usuarios.html", usuario=usuario, prestamos=prestamos_pendientes, hoy=date.today(), active_page="perfil")

@app.route("/StockMind/reportes")
@requiere_rol("administrador")
def reportes():
    if "usuario" not in session:
        return redirect(url_for("login"))
    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    total_productos = total_prestados = total_usuarios = 0
    try:
        cur = conexion.cursor()
        cur.execute("SELECT COUNT(*) FROM producto;")
        total_productos = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM prestamos WHERE estado='pendiente';")
        total_prestados = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM usuarios;")
        total_usuarios = cur.fetchone()[0] or 0

        cur.close()
    except Exception as e:
        desconectar(conexion)
        return f"Error generando reportes: {e}", 500
    finally:
        desconectar(conexion)

    return render_template(
        "reportes.html",
        total_productos=total_productos,
        total_prestados=total_prestados,
        total_usuarios=total_usuarios,
        active_page="reportes"
    )

@app.route("/StockMind/ver_reporte_excel/<tipo>")
def ver_reporte_excel(tipo):
    """Genera Excel profesional con colores, imágenes y bordes (sin logo)."""
    descargar_flag = request.args.get("descargar", "0") == "1"
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")

    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo.title()}"

    # ----- TÍTULO ----- 
    ws.merge_cells("A1:H2")
    ws["A1"] = f"StockMind - Reporte de {tipo.title()} ({datetime.now().strftime('%d/%m/%Y %H:%M')})"
    ws["A1"].font = openpyxl.styles.Font(size=16, bold=True, color="007bd4")
    ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cur = conexion.cursor()
    data_rows = []
    headers = []

    try:
        if tipo == "productos":
            query = "SELECT id_prod, nombre, cantidad, lugar, descripcion, imagen FROM producto WHERE 1=1"
            params = []
            if fecha_inicio and fecha_fin:
                query += " AND fecha_registro BETWEEN %s AND %s"
                params = [fecha_inicio + " 00:00:00", fecha_fin + " 23:59:59"]
            cur.execute(query, params)
            data_rows = cur.fetchall()
            headers = ["ID", "Nombre", "Cantidad", "Lugar", "Descripción", "Imagen"]

        elif tipo == "usuarios":
            query = "SELECT id_usu, nombre, usuario, rol, telefono, correo, foto FROM usuarios WHERE 1=1"
            params = []
            if fecha_inicio and fecha_fin:
                query += " AND fecha_registro BETWEEN %s AND %s"
                params = [fecha_inicio + " 00:00:00", fecha_fin + " 23:59:59"]
            cur.execute(query, params)
            data_rows = cur.fetchall()
            headers = ["ID", "Nombre", "Usuario", "Rol", "Teléfono", "Correo", "Foto"]

        elif tipo == "prestamos":
            query = """
                SELECT p.id_prod, p.nombre, pr.cantidad, pr.estado, pr.fecha_prestamo
                FROM prestamos pr
                LEFT JOIN producto p ON pr.id_prod = p.id_prod
                WHERE 1=1
            """
            params = []
            if fecha_inicio and fecha_fin:
                query += " AND pr.fecha_prestamo BETWEEN %s AND %s"
                params = [fecha_inicio + " 00:00:00", fecha_fin + " 23:59:59"]
            cur.execute(query, params)
            data_rows = cur.fetchall()
            headers = ["ID Producto", "Nombre Producto", "Cantidad", "Estado", "Fecha"]

        else:
            cur.close()
            desconectar(conexion)
            return "Tipo de reporte desconocido", 400

        start_row = 4
        # ----- ENCABEZADOS -----
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="007bd4")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[start_row].height = 25

        # ----- DATOS -----
        for row_idx, row in enumerate(data_rows, start_row + 1):
            fill_color = "E6F2FA" if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, value in enumerate(row, 1):
                if (tipo == "productos" and col_idx == 6) or (tipo == "usuarios" and col_idx == 7):
                    imagen_filename = value
                    if imagen_filename:
                        imagen_path = os.path.join(UPLOAD_FOLDER, imagen_filename)
                        if os.path.exists(imagen_path):
                            try:
                                img = XLImage(imagen_path)
                                img.width = 50
                                img.height = 50
                                ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")
                                ws.row_dimensions[row_idx].height = 40
                                ws.column_dimensions[get_column_letter(col_idx)].width = 12
                            except:
                                ws.cell(row=row_idx, column=col_idx, value="Sin imagen")
                        else:
                            ws.cell(row=row_idx, column=col_idx, value="Sin imagen")
                    else:
                        ws.cell(row=row_idx, column=col_idx, value="Sin imagen")
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = openpyxl.styles.PatternFill("solid", fgColor=fill_color)
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # ----- BORDES -----
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        cur.close()
    except Exception as e:
        desconectar(conexion)
        return f"Error generando Excel: {e}", 500
    finally:
        desconectar(conexion)

    # Ajustar ancho columnas automáticamente
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 5

    # ----- HISTORIAL -----
    if "id_usuario" in session:
        accion = "Descargó reporte" if descargar_flag else "Visualizó reporte"
        registrar_historial(
            session["id_usuario"],
            accion,
            f"{accion} de {tipo.title()}"
        )

    # ----- GUARDAR Y ENVIAR -----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"reporte_{tipo}_{fecha_actual}.xlsx"

    # FLAG: ver o descargar
    ver_flag = request.args.get("ver", "0") == "1"
    as_attachment_flag = not ver_flag  # si es ver, no descarga; si no, descarga

    return send_file(
        output,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=as_attachment_flag
    )


# ---------- Dashboard de préstamos ---------
@app.route("/StockMind/prestamos")
@requiere_rol("administrador", "encargado", )
def prestamos():
    if "usuario" not in session:
        return redirect(url_for("login"))

    conexion = conectar()
    cursor = conexion.cursor(cursor_factory=RealDictCursor)

    # ----- Traer préstamos existentes ordenados por ID -----
    cursor.execute("""
        SELECT pr.id_prestamo, u.nombre AS usuario, p.nombre AS producto,
               pr.cantidad, pr.fecha_prestamo, pr.fecha_devolucion, pr.estado
        FROM prestamos pr
        JOIN usuarios u ON pr.id_usu = u.id_usu
        JOIN producto p ON pr.id_prod = p.id_prod
        ORDER BY pr.id_prestamo DESC
    """)
    prestamos = cursor.fetchall()

    # ----- Traer usuarios -----
    cursor.execute("SELECT id_usu, nombre FROM usuarios")
    usuarios = cursor.fetchall()

    # ----- Traer productos y calcular disponibles -----
    cursor.execute("""
        SELECT p.id_prod, p.nombre, p.cantidad AS total,
            COALESCE(SUM(CASE WHEN pr.estado='pendiente' THEN pr.cantidad ELSE 0 END),0) AS prestados
        FROM producto p
        LEFT JOIN prestamos pr ON p.id_prod = pr.id_prod
        WHERE p.activo = TRUE
        GROUP BY p.id_prod
        ORDER BY p.id_prod
    """)
    productos = cursor.fetchall()
    # Agregar campo 'disponible'
    for p in productos:
        p['disponible'] = p['total'] - p['prestados']


    cursor.close()
    desconectar(conexion)

    return render_template(
        "prestamos.html",
        prestamos=prestamos,
        usuarios=usuarios,
        productos=productos,
        active_page="prestamos"
    )


# ---------- Registrar préstamo ----------
@app.route("/StockMind/registrar_prestamo", methods=["POST"])
def registrar_prestamo():
    try:
        id_usu = int(request.form["usuario"])
        id_prod = int(request.form["producto"])
        cantidad = int(request.form["cantidad"])
        fecha_prestamo = request.form["fecha_prestamo"]
        fecha_devolucion = request.form["fecha_devolucion"]
    except Exception as e:
        flash(f"Datos inválidos: {e}", "danger")
        return redirect(url_for("prestamos"))

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("prestamos"))

    try:
        cursor = conexion.cursor()

        # Verificar que el producto esté activo (columna boolean)
        cursor.execute("SELECT activo FROM producto WHERE id_prod=%s", (id_prod,))
        prod_row = cursor.fetchone()
        if not prod_row or prod_row[0] is not True:
            flash("El producto seleccionado está inactivo y no se puede prestar", "danger")
            cursor.close()
            desconectar(conexion)
            return redirect(url_for("prestamos"))


        # Verificar stock disponible dinámico (solo prestamos pendientes y producto activo)
        cursor.execute("""
            SELECT p.cantidad - COALESCE(SUM(pr.cantidad), 0) AS disponible
            FROM producto p
            LEFT JOIN prestamos pr ON p.id_prod = pr.id_prod AND pr.estado='pendiente'
            WHERE p.id_prod=%s AND p.activo = TRUE
            GROUP BY p.id_prod, p.cantidad
        """, (id_prod,))
        disponible_row = cursor.fetchone()
        disponible = disponible_row[0] if disponible_row else 0
        if disponible is None:
            disponible = 0

        if disponible < cantidad:
            flash(f"No hay suficiente stock disponible. Disponible: {disponible}", "danger")
            cursor.close()
            desconectar(conexion)
            return redirect(url_for("prestamos"))

        # Insertar préstamo
        cursor.execute("""
            INSERT INTO prestamos (id_usu, id_prod, cantidad, fecha_prestamo, fecha_devolucion, estado)
            VALUES (%s, %s, %s, %s, %s, 'pendiente')
            RETURNING id_prestamo
        """, (id_usu, id_prod, cantidad, fecha_prestamo, fecha_devolucion))
        id_prestamo_row = cursor.fetchone()
        id_prestamo = id_prestamo_row[0] if id_prestamo_row else None

        # Insertar en historial (registro del sistema)
        cursor.execute("""
            INSERT INTO historial (id_usu, id_prod, accion, detalle)
            VALUES (%s, %s, %s, %s)
        """, (
            id_usu,
            id_prod,
            "Registrar préstamo",
            f"Se prestaron {cantidad} unidades del producto ID {id_prod} con devolución el {fecha_devolucion}"
        ))

        conexion.commit()
        cursor.close()

        # Notificación (intento enviar)
        if id_prestamo:
            notificacion_prestamo(id_prestamo)
            flash("Préstamo registrado correctamente y notificación enviada", "success")
        else:
            flash("Préstamo registrado correctamente", "success")

    except Exception as e:
        conexion.rollback()
        flash(f"Error registrando préstamo: {e}", "danger")
    finally:
        desconectar(conexion)

    return redirect(url_for("prestamos"))


# ---------- Registrar devolución ----------

@app.route("/StockMind/devolucion", methods=["POST"])
def devolucion():
    # aceptamos varios nombres por si el formulario usa "cantidad" o "cantidad_devolver"
    id_prestamo = request.form.get("id_prestamo")
    cantidad_devolver = request.form.get("cantidad") or request.form.get("cantidad_devolver")
    descripcion = request.form.get("descripcion")

    if not id_prestamo or not cantidad_devolver:
        flash("Faltan datos para la devolución", "danger")
        return redirect(url_for("prestamos"))

    # validar entero
    try:
        cantidad_devolver = int(cantidad_devolver)
    except ValueError:
        flash("Cantidad no válida", "danger")
        return redirect(url_for("prestamos"))

    conexion = conectar()
    if not conexion:
        flash("No se pudo conectar a la base de datos", "danger")
        return redirect(url_for("prestamos"))

    try:
        cursor = conexion.cursor()

        # Traer el préstamo sin filtrar por estado
        cursor.execute("""
            SELECT id_prod, cantidad, id_usu, estado
            FROM prestamos
            WHERE id_prestamo = %s
        """, (id_prestamo,))
        prestamo = cursor.fetchone()

        print("DEBUG -> prestamo fetchone:", prestamo)  # revisa consola/logs

        if not prestamo:
            flash("El préstamo no existe", "warning")
            cursor.close()
            desconectar(conexion)
            return redirect(url_for("prestamos"))

        # unpack y forzar tipos
        id_prod, cantidad_actual, id_usu, estado_actual = prestamo
        cantidad_actual = int(cantidad_actual)
        estado_actual = (estado_actual or "").lower()

        print(f"DEBUG -> id_prestamo={id_prestamo} cantidad_actual={cantidad_actual} cantidad_devolver={cantidad_devolver} estado_actual={estado_actual}")

        if estado_actual != "pendiente":
            flash("Este préstamo ya fue cerrado", "warning")
            cursor.close()
            desconectar(conexion)
            return redirect(url_for("prestamos"))

        if cantidad_devolver <= 0 or cantidad_devolver > cantidad_actual:
            flash("Cantidad de devolución no válida", "danger")
            cursor.close()
            desconectar(conexion)
            return redirect(url_for("prestamos"))

        ahora = datetime.now()

        # Devolución total
        if cantidad_devolver == cantidad_actual:
            # Intentamos setear cantidad = 0 y estado = 'devuelto' (mejor lógica si permites cantidad=0)
            try:
                cursor.execute("""
                    UPDATE prestamos
                    SET cantidad = 0, estado = %s, fecha_devolucion = %s
                    WHERE id_prestamo = %s
                """, ('devuelto', ahora, id_prestamo))
            except Exception as e_upd:
                # si falla (p.ej. CHECK constraint que no permite 0), hacemos fallback solo cambiando el estado
                conexion.rollback()
                print("WARN -> fallo al poner cantidad=0, fallback a solo estado. error:", e_upd)
                cursor.execute("""
                    UPDATE prestamos
                    SET estado = %s, fecha_devolucion = %s
                    WHERE id_prestamo = %s
                """, ('devuelto', ahora, id_prestamo))

        else:
            # Devolución parcial: restamos la cantidad
            cursor.execute("""
                UPDATE prestamos
                SET cantidad = cantidad - %s
                WHERE id_prestamo = %s
            """, (cantidad_devolver, id_prestamo))

        # Guardar en historial (detalle con la descripción si la hay)
        detalle = f"Devolución de {cantidad_devolver} unidades del producto ID {id_prod}"
        if descripcion:
            detalle += f". Observación: {descripcion}"

        cursor.execute("""
            INSERT INTO historial (id_usu, id_prod, accion, detalle)
            VALUES (%s, %s, %s, %s)
        """, (id_usu, id_prod, "Devolución", detalle))

        conexion.commit()
        cursor.close()

        # Notificación (si existe)
        try:
            notificacion_devolucion(id_prestamo)
        except Exception as e_not:
            print("WARN -> notificacion_devolucion falló:", e_not)

        flash("Devolución registrada correctamente", "success")

    except Exception as e:
        conexion.rollback()
        print("ERROR -> fallo en devolucion():", e)
        flash(f"Error al registrar la devolución: {e}", "danger")
    finally:
        desconectar(conexion)

    return redirect(url_for("prestamos"))


# ---------- HISTORIAL ----------
@app.route("/StockMind/historial")
@requiere_rol("administrador")
def historial():
    if "id_usuario" not in session:
        return redirect(url_for("login"))

    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    cur = conexion.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT h.id_mov, u.nombre AS usuario, h.accion, h.detalle, h.fecha
            FROM historial h
            JOIN usuarios u ON h.id_usu = u.id_usu
            ORDER BY h.fecha DESC
        """)
        registros = cur.fetchall()
    except Exception as e:
        cur.close()
        desconectar(conexion)
        return f"Error consultando historial: {e}", 500
    finally:
        cur.close()
        desconectar(conexion)

    return render_template("historial.html", registros=registros, active_page="historial")


# ---------- REGISTRAR HISTORIAL ----------
def registrar_historial(id_usuario, accion, detalle=""):
    if not id_usuario:
        return
    try:
        conexion = conectar()
        if not conexion:
            print("No se pudo conectar para registrar historial")
            return
        cur = conexion.cursor()
        cur.execute("""
            INSERT INTO historial (id_usu, accion, detalle)
            VALUES (%s, %s, %s)
        """, (id_usuario, accion, detalle))
        conexion.commit()
        cur.close()
        desconectar(conexion)
    except Exception as e:
        print(f"Error registrando historial: {e}")

@app.route("/historial/filtrar", methods=["GET"])
def historial_filtrar():
    if "id_usuario" not in session:
        return redirect(url_for("login"))

    fecha_inicio = request.args.get("inicio")
    fecha_fin = request.args.get("fin")
    tipo = request.args.get("tipo", "todos")  # <- capturamos el tipo, default todos

    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    cur = conexion.cursor(cursor_factory=RealDictCursor)

    # Base de la consulta
    query = """
        SELECT h.id_mov, u.nombre AS usuario, h.accion, h.detalle, h.fecha
        FROM historial h
        JOIN usuarios u ON h.id_usu = u.id_usu
        WHERE 1=1
    """
    params = []

    # Filtro de fechas
    if fecha_inicio:
        query += " AND h.fecha >= %s"
        params.append(fecha_inicio)

    if fecha_fin:
        query += " AND h.fecha <= %s"
        params.append(fecha_fin)

    # Filtro por tipo de acción
    if tipo and tipo != "todos":
        if tipo == "productos":
            query += " AND h.accion ILIKE '%%producto%%'"
        elif tipo == "prestamos":
            query += " AND (h.accion ILIKE '%%préstamo%%' OR h.accion ILIKE '%%prestamo%%' OR h.accion ILIKE '%%devolución%%' OR h.accion ILIKE '%%devolucion%%')"
        elif tipo == "usuarios":
            query += " AND h.accion ILIKE '%%usuario%%'"
        elif tipo == "reportes":
            query += " AND h.accion ILIKE '%%reporte%%'"

    query += " ORDER BY h.fecha DESC"

    try:
        cur.execute(query, params)
        registros = cur.fetchall()
    except Exception as e:
        cur.close()
        desconectar(conexion)
        return f"Error filtrando historial: {e}", 500
    finally:
        cur.close()
        desconectar(conexion)

    return render_template("historial.html", registros=registros, tipo=tipo, active_page="historial")

@app.route("/StockMind/historial/excel")
def historial_excel():
    if "usuario" not in session:
        return redirect(url_for("login"))

    # 🔎 Capturar filtros desde la URL (GET)
    usuario_filtro = request.args.get("usuario")
    accion_filtro = request.args.get("accion")
    fecha_inicio = request.args.get("inicio")
    fecha_fin = request.args.get("fin")
    tipo_filtro = request.args.get("tipo")

    try:
        conexion = conectar()
        cur = conexion.cursor()

        query = """
            SELECT h.id_mov, u.usuario, h.accion, h.detalle, h.fecha
            FROM historial h
            LEFT JOIN usuarios u ON h.id_usu = u.id_usu
            WHERE 1=1
        """
        params = []

        # Filtros dinámicos
        if usuario_filtro:
            query += " AND u.usuario ILIKE %s"
            params.append(f"%{usuario_filtro}%")
        if accion_filtro:
            query += " AND h.accion ILIKE %s"
            params.append(f"%{accion_filtro}%")
        if fecha_inicio and fecha_fin:
            query += " AND h.fecha BETWEEN %s AND %s"
            params.append(fecha_inicio + " 00:00:00")
            params.append(fecha_fin + " 23:59:59")

        # Filtro por tipo de botón
        mapa_tipos = {
            "todos": None,
            "productos": "producto",
            "prestamos": "préstamo",
            "usuarios": "usuario",
            "reportes": "reporte"
        }
        valor_tipo = mapa_tipos.get(tipo_filtro)
        if valor_tipo:
            query += " AND h.accion ILIKE %s"
            params.append(f"%{valor_tipo}%")

        query += " ORDER BY h.fecha DESC"
        cur.execute(query, params)
        registros = cur.fetchall()
        cur.close()
        desconectar(conexion)

    except Exception as e:
        print(f"Error al generar Excel: {e}")
        return "Error generando reporte", 500

    # ---------- CREAR EXCEL HISTORIAL PROFESIONAL SIN LOGO ----------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # ----- TÍTULO -----
    nombres_tipos = {
        "todos": "Todos los movimientos",
        "productos": "Productos",
        "prestamos": "Préstamos",
        "usuarios": "Usuarios",
        "reportes": "Reportes"
    }
    titulo_filtro = nombres_tipos.get(tipo_filtro, "Historial")
    ws.merge_cells("A1:E1")
    ws["A1"] = f"StockMind - Historial de {titulo_filtro} ({datetime.now().strftime('%d/%m/%Y %H:%M')})"
    ws["A1"].font = openpyxl.styles.Font(size=14, bold=True, color="007bd4")
    ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # ----- ENCABEZADOS -----
    encabezados = ["#", "Usuario", "Acción", "Detalle", "Fecha"]
    start_row = 3
    for col_num, header in enumerate(encabezados, start=1):
        c = ws.cell(row=start_row, column=col_num, value=header)
        c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        c.fill = openpyxl.styles.PatternFill("solid", fgColor="007bd4")
        c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 25

    # ----- DATOS -----
    fila = start_row + 1
    for r in registros:
        fill_color = "E6F2FA" if fila % 2 == 0 else "FFFFFF"  # filas alternadas
        ws.cell(row=fila, column=1, value=r[0]).fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=fila, column=2, value=r[1] or "N/A").fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=fila, column=3, value=r[2]).fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=fila, column=4, value=r[3] or "").fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=fila, column=5, value=r[4].strftime("%Y-%m-%d %H:%M:%S")).fill = PatternFill("solid", fgColor=fill_color)
        for col in range(1, 6):
            ws.cell(row=fila, column=col).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        fila += 1

    # ----- BORDES -----
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # ----- AJUSTAR ANCHOS DE COLUMNAS -----
    columnas = [5, 20, 20, 100, 20]  # ancho personalizado
    for i, ancho in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ----- GUARDAR EN MEMORIA -----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"historial_{titulo_filtro}_{fecha_actual}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


#----------logout--------------------
@app.route("/StockMind/logout")
@requiere_rol("administrador", "usuario", "encargado")
def logout():
    session.clear()
    return redirect(url_for("login"))

#----------------Pagina de inico-------------------
@app.route('/StockMind/inicio')
@requiere_rol("administrador", "encargado", "usuario")
def inicio():
    if "usuario" not in session:
        return redirect(url_for("login"))
    conexion = conectar()
    if not conexion:
        return "No se pudo conectar a la base de datos", 500

    try:
        cur = conexion.cursor()

        # Total productos
        cur.execute("SELECT COUNT(*) FROM producto;")
        total_productos = cur.fetchone()[0] or 0

        # Préstamos activos (ej. estado='pendiente')
        cur.execute("SELECT COUNT(*) FROM prestamos WHERE estado = 'pendiente';")
        prestamos_activos = cur.fetchone()[0] or 0

        # Total usuarios
        cur.execute("SELECT COUNT(*) FROM usuarios;")
        total_usuarios = cur.fetchone()[0] or 0

        # Último acceso (desde sesión o BD)
        ultimo_login = session.get("ultimo_login", "Sin registros")

        cur.close()
    except Exception as e:
        desconectar(conexion)
        return f"Error cargando inicio: {e}", 500
    finally:
        desconectar(conexion)

    return render_template(
        "principal_usu.html",
        total_productos=total_productos,
        prestamos_activos=prestamos_activos,
        total_usuarios=total_usuarios,
        ultimo_login=ultimo_login,
        active_page="inicio"
    )

@app.route("/StockMind/configuracion")
def configuracion():
    if "id_usuario" not in session:
        return redirect(url_for("login"))
    return render_template("configuracion.html", active_page="configuracion")




if __name__ == "__main__":
    app.run(debug=True, port=8000)

